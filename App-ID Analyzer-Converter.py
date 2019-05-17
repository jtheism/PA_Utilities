import csv
import datetime
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from tkinter.filedialog import askopenfilename, asksaveasfilename, Tk
# from tkinter import messagebox
# Tk().withdraw()  # Prevents tkinter window from opening
root = Tk()
root.withdraw()
root.focus_force()
now = datetime.datetime.now().strftime("%m/%d/%Y")


def mode_1_import():
    """

    :return: dictionary {rulename: [set(apps), set(dest_port/protocol), hit value(T/F)]
    """
    rules_dict = None  # Eliminates Pycharm Error
    unk_tu_dict = {}  # dict for unknown tcp and udp log entries
    # parse ruleset and get rule names and details
    try:
        with open(askopenfilename(initialdir=os.path.expanduser("~/Desktop"),
                                  title="Select Security Rules CSV exported from Palo Alto firewall.")) as rules_file:
            # with open("C:\\Users\\Jason\\Desktop\\policies.csv") as rules_file:

            csv_ruleset = csv.reader(rules_file, delimiter=',')
            # Test for input
            # row = 0
            # for log_entry in csv_ruleset:
            #     print(log_entry[1], log_entry[11], f'Entry: {row}')
            #     row += 1
            rules_dict = {}
            for crow in csv_ruleset:
                # Values are:       [0], [1], [2], [4] are taken from logs, [3] from policies
                # [0]set(apps), [1]set(ports), [2]boolean of rule hit in logs,
                # [3][src zone, src add, dst zone, dest add], [4]boolean: unknown tcp or udp found
                rules_dict[crow[1]] = [set(), set(), False, [crow[4], crow[5], crow[8], crow[9]], False]
            # for k in rules_dict:
            #     print(f'key: {k}\t Value: {rules_dict[k]}')
    except FileNotFoundError:
        print("Process cancelled.")
        exit()

    # read in csv file of logs
    try:
        with open(askopenfilename(initialdir=os.path.expanduser("~/Desktop"),
                                  title="Select Log File CSV exported from Palo Alto firewall.")) as log_file:
            # with open("C:\\Users\\Jason\\Desktop\\log2.csv") as log_file:
            # parse logs to get apps and dest_ports for each rule and mark rule as hit
            csv_logs = csv.reader(log_file, delimiter=',')
            ignored_apps = ["insufficient-data", "unknown-udp", "unknown-tcp", "incomplete"]
            for lrow in csv_logs:
                # ignore these apps (incomplete, insufficient-data, unknown-udp, unknown-tcp)
                if lrow[14] not in ignored_apps:
                    try:
                        # lrow[11] is the rule name
                        rules_dict[lrow[11]][2] = True
                    except KeyError:
                        pass
                    try:
                        rules_dict[lrow[11]][0].add(lrow[14])
                    except KeyError:
                        pass
                    try:
                        # add tcp or udp and port
                        # rules_dict[lrow[11]][1].add(f'{lrow[14]}:{lrow[25]}/{lrow[29]}')
                        rules_dict[lrow[11]][1].add((lrow[14], lrow[25], lrow[29]))
                    except KeyError:
                        pass
                elif lrow[14] in ["unknown-udp", "unknown-tcp"]:
                    rules_dict[lrow[11]][4] = True
                    try:
                        rules_dict[lrow[11]][2] = True
                    except KeyError:
                        pass
                    try:
                        # rules_dict[lrow[11]][1].add(f'{lrow[14]}:{lrow[25]}/{lrow[29]}')
                        rules_dict[lrow[11]][1].add((lrow[14], lrow[25], lrow[29]))
                    except KeyError:
                        pass
                    # For unknown dict
                    # key = rule name
                    # value = list [{port and tcp/udp pair: count}, ]
                    unk_tu_dict.setdefault(lrow[11])  # creates rule name dict key
                    # print(unk_tu_dict)
                    # print(unk_tu_dict[lrow[11]])
                    if unk_tu_dict[lrow[11]] is None:
                        unk_tu_dict[lrow[11]] = {}
                        unk_tu_dict[lrow[11]].setdefault(f'{lrow[25]}/{lrow[29]}')
                        unk_tu_dict[lrow[11]][f'{lrow[25]}/{lrow[29]}'] = 1
                    else:
                        unk_tu_dict[lrow[11]].setdefault(f'{lrow[25]}/{lrow[29]}')
                        unk_tu_dict[lrow[11]][f'{lrow[25]}/{lrow[29]}'] += 1

                else:
                    # try:
                    #     rules_dict[lrow[11]][2] = True
                    # except KeyError:
                    #     pass
                    # try:
                    #     rules_dict[lrow[11]][1].add(f'{lrow[14]}:{lrow[25]}/{lrow[29]}')
                    # except KeyError:
                    #     pass
                    pass
    except FileNotFoundError:
        print("Process cancelled.")
        exit()
    # try:
    del rules_dict["Name"]
    # except KeyError:
    #     pass
    del rules_dict["intrazone-default"]
    del rules_dict["interzone-default"]
    return rules_dict, unk_tu_dict


def create_wb_out(h_dict, u_dict):
    """

    :param h_dict: h_dict produced by mode1_import function
    :param u_dict: u_dict produced by mode1_import function
    :return: workbook
    """

    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "App-ID Analysis"
    ws["A1"] = "Log Hits / Create App-ID Rule"
    ws["A1"].font = Font(bold=True, size=14, color="2f75b5")
    ws["B1"] = "Rule Name"
    ws["B1"].font = Font(bold=True, size=14, color="2f75b5")
    ws["C1"] = "Apps Identified"
    ws["C1"].font = Font(bold=True, size=14, color="2f75b5")
    ws["D1"] = "Dest.Ports Seen"
    ws["D1"].font = Font(bold=True, size=14, color="2f75b5")
    ws["E1"] = "Src.Zone"
    ws["E1"].font = Font(bold=True, size=14, color="2f75b5")
    ws["F1"] = "Src.Address"
    ws["F1"].font = Font(bold=True, size=14, color="2f75b5")
    ws["G1"] = "Dest.Zone"
    ws["G1"].font = Font(bold=True, size=14, color="2f75b5")
    ws["H1"] = "Dest.Address"
    ws["H1"].font = Font(bold=True, size=14, color="2f75b5")
    ws["I1"] = "Unknown TCP/UDP Found"
    ws["I1"].font = Font(bold=True, size=14, color="2f75b5")
    for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col].width = 25
    ws.column_dimensions["A"].width = 28
    row_num = 2
    for rule in h_dict:
        ws[f'A{row_num}'].value = h_dict[rule][2]
        ws[f'B{row_num}'].value = rule
        if str(h_dict[rule][0]) != "set()":
            ws[f'C{row_num}'].value = str(h_dict[rule][0])
        if str(h_dict[rule][1]) != "set()":
            # ws[f'D{row_num}'].value = str(h_dict[rule][1])
            temp_port_set = set()
            for port_tup in h_dict[rule][1]:
                temp_port_set.add(f'{port_tup[0]}: {port_tup[1]}/{port_tup[2]}')
                ws[f'D{row_num}'].value = str(temp_port_set)
        ws[f'E{row_num}'].value = h_dict[rule][3][0]
        ws[f'F{row_num}'].value = h_dict[rule][3][1]
        ws[f'G{row_num}'].value = h_dict[rule][3][2]
        ws[f'H{row_num}'].value = h_dict[rule][3][3]
        ws[f'I{row_num}'].value = h_dict[rule][4]
        row_num += 1

    for row in ws.iter_rows(min_row=2):
        if row[0].value:
            row[0].font = Font(color="33b743", bold=True, size=14)
            for cell in row:
                cell.fill = PatternFill(start_color="ebf1de", end_color="ebf1de", fill_type="solid")
        else:
            row[0].font = Font(color="ac0000")
    ws_unknowns = wb_out.create_sheet("Unknown TCP-UDP")
    ws_unknowns["A1"] = "Rule Name"
    ws_unknowns["A1"].font = Font(bold=True, size=14, color="2f75b5")
    ws_unknowns["B1"] = "Port"
    ws_unknowns["B1"].font = Font(bold=True, size=14, color="2f75b5")
    ws_unknowns["C1"] = "Hit Count"
    ws_unknowns["C1"].font = Font(bold=True, size=14, color="2f75b5")
    for col in ["A", "B", "C"]:
        ws_unknowns.column_dimensions[col].width = 30
    nrow = 2
    for rulename in u_dict:
        for pair in u_dict[rulename]:
            # for k in u_dict[rulename][pair]:
            ws_unknowns[f'A{nrow}'].value = rulename
            ws_unknowns[f'B{nrow}'].value = pair
            ws_unknowns[f'C{nrow}'].value = u_dict[rulename][pair]
            nrow += 1
    return wb_out


def m2_create_output(file):

    out_list = ["set tag App-ID color color23\n"]
    remove_chars = ",'{}"  # Because the apps and ports are Python sets in excel, they show as {"one", "two, "three"}
    # need to remove those characters so you just get a string of: one two three

    # wb_in = openpyxl.load_workbook(file)
    wb_sh = file["App-ID Analysis"]
    for row in wb_sh.iter_rows(min_row=2):
        if row[0].value:
            orig_name = row[1].value
            rule_name = f'{row[1].value[:51]} - App-ID'
            apps = str(row[2].value)
            for ch in remove_chars:
                apps = apps.replace(ch, "")
            out_list.append(f'copy rulebase security rules "{orig_name}" to "{rule_name}"\n')
            out_list.append(f'move rulebase security rules "{rule_name}" before "{orig_name}"\n')
            out_list.append(f'delete rulebase security rules "{rule_name}" application\n')
            out_list.append(f'delete rulebase security rules "{rule_name}" service\n')
            out_list.append(f'set rulebase security rules "{rule_name}" description "Added on {now}." application '
                            f'[ {apps} ] service application-default tag App-ID\n')
    return out_list


# Select mode (read-in logs/rules or create output)
while True:
    mode = input("Select mode:\n\t1. Read in logs and rules csv files."
                 "\n\t2. Create App-ID rules set commands output.\n")
    if mode not in ["1", "2"]:
        print("Please enter 1 or 2.")
        continue
    else:
        break

# 1:
# read in csv file of ruleset
if mode == "1":
    hits_dict, unknown_dict = mode_1_import()
    # rules_dict = mode_1_parse_rules(csv_rules)
    # Test Output
    # for r in hits_dict:
    #     print(f'Name: {r}\n\tApps: {hits_dict[r][0]}\n\tPorts: {hits_dict[r][1]}\n\tRule hit: {hits_dict[r][2]} '
    #           f'\n\t\tSrc.Zone: {hits_dict[r][3][0]}, Src.Addr: {hits_dict[r][3][1]}, '
    #           f'Dest.Zone: {hits_dict[r][3][2]}, Dest.Addr: {hits_dict[r][3][3]}')

    # Create workbook
    workbook_out = create_wb_out(hits_dict, unknown_dict)
    # Save workbook
    try:
        workbook_out.save(asksaveasfilename(parent=root, initialdir=os.path.expanduser("~/Desktop"),
                                            title="Create App-ID workbook:", defaultextension='.xlsx',
                                            filetypes=[('Excel File', '.xlsx')]))
    except FileNotFoundError:
        exit()

# 2:
# Create app-id rule via copy and move it above existing rule
if mode == "2":
    wb_file = None
    try:
        wb_file = openpyxl.load_workbook(askopenfilename(parent=root, initialdir=os.path.expanduser("~/Desktop"),
                                                         title="Re-Import App-ID Analysis workbook..."))
    except:
        print("Process cancelled.")
        exit()

    set_cmds = m2_create_output(wb_file)

    try:
        with open(asksaveasfilename(parent=root, initialdir=os.path.expanduser("~/Desktop"),
                                    title="Create set command output file:", defaultextension='.txt',
                                    filetypes=[('Text File', '.txt')]), "w+") as output:
            for cmd in set_cmds:
                output.write(cmd)
    except FileNotFoundError:
        print("Process cancelled.")
        exit()
